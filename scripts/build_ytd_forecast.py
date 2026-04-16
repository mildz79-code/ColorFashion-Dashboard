"""
Build 2026 YTD Actuals + Forecast spreadsheet for Color Fashion Dye & Finishing.

Structure:
  Col A: Category
  Cols B–D:   Jan, Feb, Mar  (ACTUALS, yellow-tinted header)
  Cols E–M:   Apr–Dec         (FORECAST, gray-tinted header)
  Col N: YTD Actuals (Q1)
  Col O: FY Forecast (YTD Actuals + Apr–Dec Forecast)
  Col P: Original Budget
  Col Q: Variance vs Budget (FY Forecast - Budget)

Forecast blend (Apr–Dec per line item):
  = AVERAGE(Q1 monthly average, Original Budget monthly)
  = ( (B+C+D)/3 + Budget/1 ) / 2
  Where rows the original budget didn't include, we use just the Q1 avg.
"""
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import os
# Resolve paths relative to the repo root (parent of the scripts/ folder)
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(REPO_ROOT, 'data', '2026')
BUDGET_FILE = os.path.join(REPO_ROOT, '2026_monthly_budget.xlsx')
OUT = os.path.join(DATA_DIR, '2026_YTD_Forecast.xlsx')

# ---------- 1. Parse actuals from 3 P&L CSVs ----------
def read_pl(path):
    items = {}
    with open(path) as f:
        for row in csv.reader(f):
            if len(row) < 2 or row[1] in ("", None):
                continue
            try:
                items[row[0]] = float(row[1])
            except (TypeError, ValueError):
                pass
    return items

jan_pl = read_pl(os.path.join(DATA_DIR, 'Jan_PL.csv'))
feb_pl = read_pl(os.path.join(DATA_DIR, 'Feb_PL.csv'))
mar_pl = read_pl(os.path.join(DATA_DIR, 'Mar_PL.csv'))

# ---------- 2. Load original budget baseline (Jan column values only) ----------
wb_budget = load_workbook(BUDGET_FILE, data_only=True)
ws_budget = wb_budget['2026 Monthly Budget']
budget_lookup = {}
for r in range(1, ws_budget.max_row + 1):
    label = ws_budget.cell(row=r, column=1).value
    val = ws_budget.cell(row=r, column=2).value  # col B = Jan baseline
    if label and isinstance(val, (int, float)):
        budget_lookup[label] = val

# ---------- 3. Define the row layout ----------
# Each entry: (label, source_keys_or_None, type)
# type: 'hdr' section header, 'data' regular line, 'sub' subtotal, 'calc' formula row
ROWS = [
    ('REVENUE', None, 'hdr'),
    ('Total Revenue (Sales)',              ['Sales'], 'data'),
    (None, None, 'blank'),
    ('COST OF GOODS SOLD (COGS)', None, 'hdr'),
    ('Direct Labor - Samuel Hale',         ['Direct Labor- Samuel Hale'], 'data'),
    ('Direct Labor - Workforce',           ['Direct Labor- Workforce'], 'data'),
    ('Payroll - Other',                    ['Payroll - Other'], 'data'),
    ('Chemical & Dyestuffs',               ['Chemical & Dyestuffs'], 'data'),
    ('Finishing Supplies - Paper Tube',    ['Paper Tube'], 'data'),
    ('Finishing Supplies - Poly Bags',     ['Poly Bags'], 'data'),
    ('Lab Supplies (Testing)',             ['Lab Supplies (Testing)'], 'data'),
    ('Plant Supplies & Parts',             ['Plant Supplies & Parts'], 'data'),
    ('Freight and Shipping Costs',         ['Freight and Shipping Costs'], 'data'),
    ('Truck Repair',                       ['Truck-Repair'], 'data'),
    ('Insurance - Liability & Property',   ['Insurance-Liability & Property'], 'data'),
    ('Utilities - Electricity',            ['Electricity'], 'data'),
    ('Utilities - Gas',                    ['Gas'], 'data'),
    ('Utilities - Water',                  ['Water'], 'data'),
    ('Utilities - Wastewater',             ['Wastewater Surcharge (Current)'], 'data'),
    ('Knitting',                           ['Knitting'], 'data'),
    ('Total COGS', None, 'sub_cogs'),
    (None, None, 'blank'),
    ('GROSS PROFIT', None, 'gross'),
    (None, None, 'blank'),
    ('OPERATING EXPENSES', None, 'hdr'),
    ('Payroll Expenses (Admin)',           ['Payroll Expenses'], 'data'),
    ('Payroll Taxes',                      ['Payroll Taxes', 'EDD'], 'data'),
    ('Employee Benefits',                  ['Employee Benefits'], 'data'),
    ('Rent Expense',                       ['Rent Expense'], 'data'),
    ('Rent Management Fee',                ['Rent Management Fee'], 'data'),
    ('Professional Fees - Legal',          ['Legal Fees'], 'data'),
    ('Professional Fees - Trucking',       ['Trucking Transportation'], 'data'),
    ('Professional Fees - Other',          ['Professional Fees - Other'], 'data'),
    ('Sales Commission',                   ['sales commision'], 'data'),
    ('Sales Promotion',                    ['Sales Promotion'], 'data'),
    ('Office Expense',                     ['Office Expense'], 'data'),
    ('Office Supplies & Printing',         ['Office Supplies & Printing'], 'data'),
    ('Computer and Internet',              ['Computer and Internet Expenses'], 'data'),
    ('Telephone Expense',                  ['Telephone Expense'], 'data'),
    ('Automobile Expense',                 ['Automobile Expense'], 'data'),
    ('Repairs - Computer',                 ['Computer'], 'data'),
    ('Repairs - Equipment',                ['Equipment'], 'data'),
    ('Insurance - Health',                 ['Insurance-Health'], 'data'),
    ('Insurance - Truck',                  ['Insurance-Truck'], 'data'),
    ('Insurance - Trade Credit',           ['Insurance-Trade Credit'], 'data'),
    ('Licenses - Permits',                 ['Permits', 'Licenses and Permits - Other'], 'data'),
    ('Post Office Charge',                 ['Post Office Charge'], 'data'),
    ('Equipment Rental',                   ['Equipment Rental'], 'data'),
    ('Contract Labor',                     ['Contract Labor'], 'data'),
    ('Outside Service',                    ['Outside Service'], 'data'),
    ('Bank Service Charges',               ['Bank Service Charges'], 'data'),
    ('Interest Expenses',                  ['Interest Expenses'], 'data'),
    ('Travel',                             ['Travel'], 'data'),
    ('Waste Disposal',                     ['Waste Disposal'], 'data'),
    ('Donations',                          ['Donations'], 'data'),
    ('Medical Expense & Supplies',         ['Medical Expense & Supplies'], 'data'),
    ('Total Operating Expenses', None, 'sub_opex'),
    (None, None, 'blank'),
    ('NET ORDINARY INCOME', None, 'net_ord'),
    ('Other Income (Refunds)',             ['Refund'], 'data'),
    ('NET INCOME (Final)', None, 'net_final'),
]

# Mapping from budget-file labels to our new rows, for pulling original budget baseline.
# If a row isn't in the budget file, the baseline is None (forecast will use Q1 avg alone).
BUDGET_LABEL_MAP = {
    'Total Revenue (Sales)':              'Total Revenue (Sales)',
    'Direct Labor - Samuel Hale':         'Direct Labor - Samuel Hale',
    'Direct Labor - Workforce':           'Direct Labor - Workforce',
    'Payroll - Other':                    'Payroll - Other',
    'Chemical & Dyestuffs':               'Chemical & Dyestuffs',
    'Finishing Supplies - Paper Tube':    'Finishing Supplies - Paper Tube',
    'Lab Supplies (Testing)':             'Lab Supplies (Testing)',
    'Plant Supplies & Parts':             'Plant Supplies & Parts',
    'Freight and Shipping Costs':         'Freight and Shipping Costs',
    'Truck Repair':                       'Truck Repair',
    'Insurance - Liability & Property':   'Insurance - Liability & Property',
    'Utilities - Gas':                    'Utilities - Gas',
    'Utilities - Water':                  'Utilities - Water',
    'Payroll Expenses (Admin)':           'Payroll Expenses (Admin)',
    'Payroll Taxes':                      'Payroll Taxes',
    'Employee Benefits':                  'Employee Benefits',
    'Rent Expense':                       'Rent Expense',
    'Rent Management Fee':                'Rent Management Fee',
    'Professional Fees - Legal':          'Professional Fees - Legal',
    'Professional Fees - Other':          'Professional Fees - Other',
    'Sales Commission':                   'Sales Commission',
    'Sales Promotion':                    'Sales Promotion',
    'Office Expense':                     'Office Expense',
    'Office Supplies & Printing':         'Office Supplies & Printing',
    'Computer and Internet':              'Computer and Internet',
    'Telephone Expense':                  'Telephone Expense',
    'Automobile Expense':                 'Automobile Expense',
    'Repairs - Computer':                 'Repairs - Computer',
    'Repairs - Equipment':                'Repairs - Equipment',
    'Insurance - Health':                 'Insurance - Health',
    'Insurance - Truck':                  'Insurance - Truck',
    'Licenses - Permits':                 'Licenses - Permits',
    'Post Office Charge':                 'Post Office Charge',
    'Equipment Rental':                   'Equipment Rental',
    'Contract Labor':                     'Contract Labor',
    'Outside Service':                    'Outside Service',
    'Bank Service Charges':               'Bank Service Charges',
    'Other Income (Refunds)':             'Other Income (Refunds)',
}

def amt(src, keys):
    return sum(src.get(k, 0) for k in keys)

# ---------- 4. Build workbook ----------
wb = Workbook()
ws = wb.active
ws.title = '2026 YTD + Forecast'

# Styles
BLUE      = PatternFill('solid', start_color='FF4472C4')
LIGHT_BLUE= PatternFill('solid', start_color='FFD9E1F2')
GREEN     = PatternFill('solid', start_color='FF92D050')
DK_GREEN  = PatternFill('solid', start_color='FF00B050')
YELLOW    = PatternFill('solid', start_color='FFFFF2CC')  # actuals columns
GRAY      = PatternFill('solid', start_color='FFE7E6E6')  # forecast columns
LIGHT_GRAY= PatternFill('solid', start_color='FFF2F2F2')

FONT_TITLE  = Font(name='Cambria', size=14, bold=True)
FONT_SUB    = Font(name='Cambria', size=10, italic=True)
FONT_HDR_W  = Font(name='Cambria', size=11, bold=True, color='FFFFFFFF')
FONT_HDR_BK = Font(name='Cambria', size=10, bold=True)
FONT_NET_W  = Font(name='Cambria', size=11, bold=True, color='FFFFFFFF')
FONT_BODY   = Font(name='Cambria', size=10)

CENTER = Alignment(horizontal='center', vertical='center')
LEFT   = Alignment(horizontal='left', vertical='center')
RIGHT  = Alignment(horizontal='right', vertical='center')

thin = Side(border_style='thin', color='FFBFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CURR_FMT = '"$"#,##0_);("$"#,##0);"-"'
PCT_FMT  = '0.0%;(0.0%);"-"'

# Title rows
ws.merge_cells('A1:Q1')
ws['A1'] = 'COLOR FASHION DYE & FINISHING — 2026 YTD Actuals + Forecast'
ws['A1'].font = FONT_TITLE
ws['A1'].alignment = CENTER

ws.merge_cells('A2:Q2')
ws['A2'] = 'Q1 Actuals (Jan–Mar) | Apr–Dec Forecast (blend: Q1 run-rate + Original Budget)'
ws['A2'].font = FONT_SUB
ws['A2'].alignment = CENTER

# Header row (row 4)
HEADER_ROW = 4
FIRST_DATA_ROW = 5
months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
headers = ['Category'] + months + ['YTD Actuals','FY Forecast','Budget','Variance vs Budget']

for i, h in enumerate(headers, start=1):
    c = ws.cell(row=HEADER_ROW, column=i, value=h)
    c.font = FONT_HDR_W
    c.fill = BLUE
    c.alignment = CENTER
    c.border = BORDER

# Sub-header row (row 3): mark actuals vs forecast
ws.cell(row=3, column=1, value='').fill = LIGHT_GRAY
for i in range(2, 5):  # cols B-D = Jan,Feb,Mar
    c = ws.cell(row=3, column=i, value='← ACTUALS →' if i == 3 else '')
    c.fill = YELLOW
    c.font = Font(name='Cambria', size=9, bold=True, italic=True)
    c.alignment = CENTER
for i in range(5, 14):  # cols E-M = Apr-Dec
    c = ws.cell(row=3, column=i, value='← FORECAST →' if i == 9 else '')
    c.fill = GRAY
    c.font = Font(name='Cambria', size=9, bold=True, italic=True)
    c.alignment = CENTER
for i in range(14, 18):  # summary cols
    c = ws.cell(row=3, column=i, value='')
    c.fill = LIGHT_GRAY

# ---------- 5. Write data rows ----------
# Column letters
COL_JAN, COL_FEB, COL_MAR = 'B', 'C', 'D'
COL_APR, COL_DEC = 'E', 'M'
COL_YTD, COL_FY, COL_BUD, COL_VAR = 'N', 'O', 'P', 'Q'

# Keep track of row numbers by label so subtotals reference them
row_of = {}
cogs_data_rows = []     # data rows within COGS
opex_data_rows = []     # data rows within OpEx

current_row = FIRST_DATA_ROW
in_cogs = False
in_opex = False
net_ord_row = None

for spec in ROWS:
    label, keys, rtype = spec

    if rtype == 'blank':
        current_row += 1
        continue

    if rtype == 'hdr':
        c = ws.cell(row=current_row, column=1, value=label)
        c.font = FONT_HDR_W
        c.fill = BLUE
        c.alignment = LEFT
        for col in range(2, 18):
            ws.cell(row=current_row, column=col).fill = BLUE
        if label.startswith('COST OF GOODS'):
            in_cogs = True; in_opex = False
        elif label.startswith('OPERATING'):
            in_opex = True; in_cogs = False
        current_row += 1
        continue

    if rtype == 'data':
        row_of[label] = current_row
        # Category
        c = ws.cell(row=current_row, column=1, value=label)
        c.font = FONT_BODY; c.alignment = LEFT

        # Jan, Feb, Mar actuals (hardcoded from CSVs)
        actuals = [amt(jan_pl, keys), amt(feb_pl, keys), amt(mar_pl, keys)]
        for i, v in enumerate(actuals):
            cc = ws.cell(row=current_row, column=2 + i, value=v)
            cc.number_format = CURR_FMT
            cc.font = FONT_BODY
            cc.fill = YELLOW

        # Budget baseline (col P) — pulled from original budget file
        budget_label = BUDGET_LABEL_MAP.get(label)
        monthly_bud = budget_lookup.get(budget_label) if budget_label else None
        annual_bud = monthly_bud * 12 if monthly_bud is not None else 0
        cb = ws.cell(row=current_row, column=16, value=annual_bud)
        cb.number_format = CURR_FMT
        cb.font = FONT_BODY
        cb.fill = LIGHT_GRAY

        # Apr–Dec forecast formula
        # = IF monthly budget exists: AVERAGE(Q1 avg, monthly budget)
        # = else: Q1 avg (average of Jan, Feb, Mar)
        q1_avg_ref = f'AVERAGE({COL_JAN}{current_row}:{COL_MAR}{current_row})'
        if monthly_bud is not None and monthly_bud != 0:
            bud_monthly_ref = f'({COL_BUD}{current_row}/12)'
            forecast_formula = f'=AVERAGE({q1_avg_ref},{bud_monthly_ref})'
        else:
            forecast_formula = f'={q1_avg_ref}'
        for col_idx in range(5, 14):  # E..M
            cf = ws.cell(row=current_row, column=col_idx, value=forecast_formula)
            cf.number_format = CURR_FMT
            cf.font = FONT_BODY
            cf.fill = GRAY

        # YTD Actuals (Q1 sum)
        ws.cell(row=current_row, column=14,
                value=f'=SUM({COL_JAN}{current_row}:{COL_MAR}{current_row})'
               ).number_format = CURR_FMT
        # FY Forecast = YTD + Apr..Dec
        ws.cell(row=current_row, column=15,
                value=f'=SUM({COL_JAN}{current_row}:{COL_DEC}{current_row})'
               ).number_format = CURR_FMT
        # Variance vs Budget = FY Forecast - Budget
        ws.cell(row=current_row, column=17,
                value=f'={COL_FY}{current_row}-{COL_BUD}{current_row}'
               ).number_format = CURR_FMT
        # Style summary cells
        for col_idx in [14, 15, 17]:
            ws.cell(row=current_row, column=col_idx).font = FONT_BODY
            ws.cell(row=current_row, column=col_idx).fill = LIGHT_GRAY
        ws.cell(row=current_row, column=14).fill = PatternFill('solid', start_color='FFFEF9E7')
        ws.cell(row=current_row, column=15).fill = PatternFill('solid', start_color='FFEAF4FC')

        if in_cogs:
            cogs_data_rows.append(current_row)
        elif in_opex:
            opex_data_rows.append(current_row)

        current_row += 1
        continue

    if rtype == 'sub_cogs':
        row_of['Total COGS'] = current_row
        c = ws.cell(row=current_row, column=1, value=label)
        c.font = FONT_HDR_BK; c.fill = LIGHT_BLUE
        for col_idx in range(2, 18):
            col_letter = get_column_letter(col_idx)
            refs = ','.join(f'{col_letter}{r}' for r in cogs_data_rows)
            f = f'=SUM({refs})'
            cc = ws.cell(row=current_row, column=col_idx, value=f)
            cc.number_format = CURR_FMT
            cc.font = FONT_HDR_BK
            cc.fill = LIGHT_BLUE
        current_row += 1
        in_cogs = False
        continue

    if rtype == 'sub_opex':
        row_of['Total Operating Expenses'] = current_row
        c = ws.cell(row=current_row, column=1, value=label)
        c.font = FONT_HDR_BK; c.fill = LIGHT_BLUE
        for col_idx in range(2, 18):
            col_letter = get_column_letter(col_idx)
            refs = ','.join(f'{col_letter}{r}' for r in opex_data_rows)
            f = f'=SUM({refs})'
            cc = ws.cell(row=current_row, column=col_idx, value=f)
            cc.number_format = CURR_FMT
            cc.font = FONT_HDR_BK
            cc.fill = LIGHT_BLUE
        current_row += 1
        in_opex = False
        continue

    if rtype == 'gross':
        row_of['GROSS PROFIT'] = current_row
        rev_row = row_of['Total Revenue (Sales)']
        cogs_row = row_of['Total COGS']
        c = ws.cell(row=current_row, column=1, value=label)
        c.font = Font(name='Cambria', size=11, bold=True)
        c.fill = GREEN
        for col_idx in range(2, 18):
            col_letter = get_column_letter(col_idx)
            f = f'={col_letter}{rev_row}-{col_letter}{cogs_row}'
            cc = ws.cell(row=current_row, column=col_idx, value=f)
            cc.number_format = CURR_FMT
            cc.font = Font(name='Cambria', size=11, bold=True)
            cc.fill = GREEN
        current_row += 1
        continue

    if rtype == 'net_ord':
        row_of['NET ORDINARY INCOME'] = current_row
        net_ord_row = current_row
        gp_row = row_of['GROSS PROFIT']
        opex_row = row_of['Total Operating Expenses']
        c = ws.cell(row=current_row, column=1, value=label)
        c.font = FONT_NET_W; c.fill = DK_GREEN
        for col_idx in range(2, 18):
            col_letter = get_column_letter(col_idx)
            f = f'={col_letter}{gp_row}-{col_letter}{opex_row}'
            cc = ws.cell(row=current_row, column=col_idx, value=f)
            cc.number_format = CURR_FMT
            cc.font = FONT_NET_W
            cc.fill = DK_GREEN
        current_row += 1
        continue

    if rtype == 'net_final':
        row_of['NET INCOME (Final)'] = current_row
        other_row = row_of['Other Income (Refunds)']
        c = ws.cell(row=current_row, column=1, value=label)
        c.font = FONT_NET_W; c.fill = DK_GREEN
        for col_idx in range(2, 18):
            col_letter = get_column_letter(col_idx)
            f = f'={col_letter}{net_ord_row}+{col_letter}{other_row}'
            cc = ws.cell(row=current_row, column=col_idx, value=f)
            cc.number_format = CURR_FMT
            cc.font = FONT_NET_W
            cc.fill = DK_GREEN
        current_row += 1
        continue

# ---------- 6. KEY METRICS section ----------
current_row += 2
metrics_header_row = current_row
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=17)
c = ws.cell(row=current_row, column=1, value='KEY METRICS')
c.font = Font(name='Cambria', size=12, bold=True, color='FFFFFFFF')
c.fill = BLUE
c.alignment = LEFT
current_row += 1

rev_row   = row_of['Total Revenue (Sales)']
gp_row    = row_of['GROSS PROFIT']
netord_row= row_of['NET ORDINARY INCOME']
netfin_row= row_of['NET INCOME (Final)']

metric_rows = [
    ('Gross Margin %',     gp_row),
    ('Operating Margin %', netord_row),
    ('Net Margin %',       netfin_row),
]
for label, ref_row in metric_rows:
    ws.cell(row=current_row, column=1, value=label).font = FONT_BODY
    for col_idx in range(2, 18):
        col_letter = get_column_letter(col_idx)
        f = f'=IFERROR({col_letter}{ref_row}/{col_letter}{rev_row},0)'
        cc = ws.cell(row=current_row, column=col_idx, value=f)
        cc.number_format = PCT_FMT
        cc.font = FONT_BODY
    current_row += 1

# ---------- 7. Column widths, freeze panes ----------
ws.column_dimensions['A'].width = 34
for col in range(2, 14):
    ws.column_dimensions[get_column_letter(col)].width = 11
ws.column_dimensions['N'].width = 13
ws.column_dimensions['O'].width = 13
ws.column_dimensions['P'].width = 13
ws.column_dimensions['Q'].width = 16
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 16

ws.freeze_panes = 'B5'

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Rows written: up to row {current_row-1}")
